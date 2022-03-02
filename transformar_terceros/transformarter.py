import os

try:
	import openpyxl
	import re
	from datetime import datetime
except ImportError:
# Ejecutamos el instalador de dependencias
	os.system("python -m pip install -r requeriments.txt")
	print("************************************************************")
	print("* Se han instalado las dependencias reinicie la aplicación *")
	print("************************************************************")
	exit()


def protect_nonetype(cell):
	if cell is None:
		return ""
	else:
		return cell

'****************************************************'
'**** Obtiene la lista de archivos a transformar ****'
'****************************************************'
def lista_archivos():
	'Recupera todos los archivos del directorio actual'
	archivos = os.listdir(".")

	'Filtramos sólo los archivos excel'
	archivos_filter = []
	for a in archivos:
		if a.endswith(".xlsx") or a.endswith(".xls"):
			archivos_filter.append(a)

	'Devolvemos la lista de archivos excel'
	return archivos_filter


'*******************************'
'**** Validadores de campos ****'
'*******************************'
def validar_persona(nifcif, persona, nombre, apellido1, apellido2, denominacion):

	nifRegex = "^\d{8}[a-zA-Z]{1}$"
	nieRegex = "^[XxTtYyZz]{1}[0-9]{7}[a-zA-Z]{1}$"
	cifRegex = "^[a-zA-Z]{1}\d{7}[a-zA-Z0-9]{1}$"

	if (nifcif == ''):
		raise NameError('Identificador es vacío')
	
	if (persona == 'F' and (nombre == '' or apellido1 == '')):
		raise NameError('Nombre de persona física incompleto')
			
	if (persona == 'J' and denominacion == ''):
		raise NameError('Nombre de persona jusrídica incompleto')

	if (persona == 'J' and re.search(cifRegex, nifcif) is None):
		raise NameError('Cif incorrecto')
	
	if (persona == 'F' and re.search(nifRegex, nifcif) is None and re.search(nieRegex, nifcif) is None):
		raise NameError('Nif o Nie incorrecto')

def validar_direccion(direccion, cp, provincia, municipio):

	if (direccion == '' or cp == '' or provincia == '' or municipio == ''):
		raise NameError('Alguna parte de la dirección es vacía')


'**********************************************************'
'**** Transforma el archivo excel pasado por parámetro ****'
'**********************************************************'
def transformar_archivo(archivo):

	'Abrimos el archivo excel origen'
	wb = openpyxl.load_workbook(archivo)
	sheet = wb.active

	'Abrimos el archivo log'
	log_file = open(os.path.splitext(os.path.basename(archivo))[0] + ".log", mode='w')
	log_file.write('Comienzo del proceso: ' + str(datetime.now()) + '\n')

	lineas_ok = 0
	lineas_error = 0

	'Creamos el wb de salida para errores'
	out_wb_errores = openpyxl.Workbook()
	out_ws_errores = out_wb_errores.active
	out_ws_errores.title = "Errores"

	'Creamos el wb de salida'
	out_wb = openpyxl.Workbook()
	out_ws = out_wb.active
	out_ws.title = "Ciudadanos"

	'Insertamos las cabeceras en wb salida'
	out_ws.append([
	'NIFCIF',
	'Personalidad',
	'Nombre',
	'Apellido1',
	'Apellido2',
	'FecNacimiento',
	'Denominacion',
	'TipoVia',
	'NombreVia',
	'NumVia',
	'EscPisPue',
	'EntidadMenor',
	'Direccion',
	'CodPostal',
	'DesProvincia',
	'DesMunicipio',	
	'Observaciones',
	'FechaAlta',
	'Telefono1',
	'Telefono2',
	'Telefono3',
	'CorreoElectronico1',	
	'CorreoElectronico2',	
	'CorreoElectronico3'])

	'Insertamos las cabeceras en wb errores'
	out_ws_errores.append([
	'Tipo de identificador',
	'Identificador',	
	'Nombre/Razón social',
	'Apellido 1	',
	'Apellido 2	',
	'Pais	',
	'Municipio	',
	'Provincia	',
	'Dirección	',
	'Código postal'])

	'** Recorremos las filas del fichero'
	for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):

		'Campo nifcif y tipo de persona'
		nifcif = str(protect_nonetype(row[1]))

		if (protect_nonetype(row[0]) == 'N' or protect_nonetype(row[0]) == 'E'):
			persona = 'F'
		elif (protect_nonetype(row[0]) == 'C' or protect_nonetype(row[0]) == 'P'):
			persona = 'J'
		else:
			out_ws_errores.append(row + tuple(["Error en tipo de persona"]))
			lineas_error += 1
			continue

		'Nombre, apellido1, apellido2'
		if (persona == 'F'):
			nombre = protect_nonetype(row[2])
			apellido1 = protect_nonetype(row[3])
			apellido2 = protect_nonetype(row[4])
			denominacion = nombre + " " + apellido1 + " " + apellido2
		else:
			nombre = ''
			apellido1 = ''	
			apellido2 = ''
			denominacion = protect_nonetype(row[2])

		'Validamos nif y persona'
		try:
			validar_persona(nifcif, persona, nombre, apellido1, apellido2, denominacion)
		except:
			out_ws_errores.append(row + tuple(["Error validación identificador o nombre/denominacion"]))
			lineas_error += 1
			continue

		'direccion y cp'
		direccion = protect_nonetype(row[8])
		cp = protect_nonetype(row[9])
		
		'decodificamos provincia'
		codigo_temp_provincia = str(protect_nonetype(row[7])).zfill(2)
		try:
			provincia = provincias[codigo_temp_provincia]
		except:
			out_ws_errores.append(row + tuple(["Error decodificando provincia"]))
			lineas_error += 1
			continue

		'decodificamos municipio'
		codigo_temp_municipio = codigo_temp_provincia + str(protect_nonetype(row[6])).zfill(4)
		try:
			municipio = municipios[codigo_temp_municipio]
		except:
			print(codigo_temp_municipio)
			out_ws_errores.append(row + tuple(["Error decodificando municipio"]))
			lineas_error += 1
			continue

		'decodificamos codigo postal'
		try:
			if (cp == '' and municipio != ''):
				cp = cpostales[codigo_temp_municipio[:-1]]
		except:
			out_ws_errores.append(row + tuple(["Error decodificando codigo postal"]))
			lineas_error += 1
			continue

		'Validamos direccion'
		try:
			validar_direccion(direccion, cp, provincia, municipio)
		except:
			out_ws_errores.append(row + tuple(["Error validando dirección"]))
			lineas_error += 1
			continue

		'Formamos la fila para añadir'
		out_ws.append([
			nifcif,
			persona,
			nombre,
			apellido1,
			apellido2,
			'',
			denominacion,
			'',
			'',
			'',
			'',
			'',
			direccion,
			cp,
			provincia,
			municipio,	
			'',
			'',
			'',
			'',
			'',
			'',	
			'',	
			''])

		lineas_ok += 1


	'** Guardamos el fichero de salida'
	nombre_fout = os.path.splitext(os.path.basename(a))[0] + "_OUT" + os.path.splitext(os.path.basename(a))[1]
	out_wb.save(nombre_fout)

	'** Guardamos el excel de errores'
	if (lineas_error>0):
		nombre_fout_errores = os.path.splitext(os.path.basename(a))[0] + "_ERRORES" + os.path.splitext(os.path.basename(a))[1]
		out_wb_errores.save(nombre_fout_errores)
	else:
		out_wb_errores.close()


	'** Finalizamos log'
	log_file.write('\n')
	log_file.write('Fin del proceso: ' + str(datetime.now()) + '\n')
	log_file.write('Terceros correctos  : ' + str(lineas_ok) + '\n')
	log_file.write('Terceros incorrectos: ' + str(lineas_error)+ '\n')
	log_file.close()
	
'*************************************'
'**** Cuerpo principal del script ****'
'*************************************'
if __name__ == '__main__':

	# Cargamos las provincias
	provincias = {}
	with open('provincias.csv', mode='r') as csv_file:
		for row in csv_file:
			r = row.split(';')
			provincias[r[0]] = r[1][:-1]

	# Cargamos los municipios
	municipios = {}
	with open('municipios.csv', mode='r') as csv_file:
		for row in csv_file:
			r = row.split(';')
			municipios[r[0]] = r[1][:-1]

	# Cargamos los codigos postales
	cpostales = {}
	with open('cpostales.csv', mode='r') as csv_file:
		for row in csv_file:
			r = row.split(';')
			cpostales[r[0]] = r[1][:-1]

	# Obtenemos la lista de archivos a transformar
	archivos = lista_archivos()

	# Iteramos transformando archivos
	for a in archivos:
		transformar_archivo(a)

